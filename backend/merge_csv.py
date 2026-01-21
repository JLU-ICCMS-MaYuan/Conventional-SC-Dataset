import csv
import json
import re
import sys
import io
from pathlib import Path
from datetime import datetime
from pydantic import ValidationError
import openpyxl

# 导入验证模型
try:
    from backend import schemas
except ImportError:
    import schemas

# 配置映射 (用于支持缩写输入)
ARTICLE_TYPE_MAP = {
    't': 'theoretical', 'theoretical': 'theoretical',
    'e': 'experimental', 'experimental': 'experimental'
}

SC_TYPE_MAP = {
    'c': 'cuprate', 'cuprate': 'cuprate',
    'i': 'iron_based', 'iron_based': 'iron_based',
    'n': 'nickel_based', 'nickel_based': 'nickel_based',
    'h': 'hydride', 'hydride': 'hydride',
    'cb': 'carbon', 'carbon': 'carbon',
    'or': 'organic', 'organic': 'organic',
    'ot': 'others', 'others': 'others'
}

CSV_EXAMPLE_TEXT = """
💡 快速上传文件格式说明:
支持格式: .xlsx, .csv, .txt (逗号分隔)
列顺序: doi, 文章类型, 超导类型, 化学式, 晶体结构, [P1, Tc1, L1, W1, N1], [P2, Tc2, L2, W2, N2]...
---------------------------------------------------------------------------------------
示例数据:
10.1038/nature14964, e, h, H3S, Im-3m, 150, 203, 2.1, 0.13, 0.5
(简写说明: e=experimental, t=theoretical | c=cuprate, i=iron_based, n=nickel_based, h=hydride, cb=carbon, or=organic, ot=others)
---------------------------------------------------------------------------------------
提示:
- 文章类型: t (理论), e (实验)
- 超导类型: c (铜基), i (铁基), n (镍基), h (氢化物), cb (碳基), or (有机), ot (其他)
- 物理数据组: 每 5 列为一组 (pressure, tc, lambda, omega_log, n_ef)，没有的数据请留空。
"""

def create_example_xlsx(file_path):
    """创建一个示例 XLSX 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Upload Example"
    
    header = [
        "doi", "article_type", "superconductor_type", "chemical_formula", "crystal_structure", 
        "pressure1", "tc1", "lambda1", "omega_log1", "n_ef1",
        "pressure2", "tc2", "lambda2", "omega_log2", "n_ef2", "更多pressure, tc, lambda, omega_log, n_ef"
    ]
    ws.append(header)
    
    row1 = ["💡 快速上传文件格式说明:"]
    ws.append(row1)
    row2 = ["支持格式: .xlsx, .csv, .txt (逗号分隔)"]
    ws.append(row2)
    row3 = ["列顺序: doi, 文章类型, 超导类型, 化学式, 晶体结构, P1, Tc1, L1, W1, N1, P2, Tc2, L2, W2, N2,..."]
    ws.append(row3)
    row4 = ["示例数据:"]
    ws.append(row4)
    row5 = ["10.1038/nature14964", "e", "h", "H3S", "Im-3m", "150", "203", "2.1", "", "0.5"]
    ws.append(row5)
    row6 = ["10.1103/PhysRevLett.122.027001", "e", "h", "LaH10", "Fm-3m", "170", "250", "", "", "", "180", "260", "1.3", "0.9", ""]
    ws.append(row6)
    row7 = ["(简写说明: e=experimental, t=theoretical | c=cuprate, i=iron_based, n=nickel_based, h=hydride, cb=carbon, or=organic, ot=others)"]
    ws.append(row7)
    row8 = ["---------------------------------------------------------------------------------------"]
    ws.append(row8)
    row9 = ["- 文章类型: t (理论), e (实验)"]
    ws.append(row9)
    row10 = ["- 超导类型: c (铜基), i (铁基), n (镍基), h (氢化物), cb (碳基), or (有机), ot (其他)"]
    ws.append(row10)
    row11 = ["- 物理数据组: 每 5 列为一组 (pressure, tc, lambda, omega_log, n_ef)，可上传多组，没有的数据请留空。"]
    ws.append(row11)


    wb.save(file_path)
    print(f"📝 已生成示例文件: {file_path}")

def extract_elements(formula):
    if not formula: return []
    symbols = re.findall(r'[A-Z][a-z]?', str(formula))
    return sorted(list(set(symbols)))

def to_float(val):
    if val is None or str(val).strip() == "": return None
    try:
        return float(val)
    except ValueError:
        return None

def parse_rows(rows):
    """
    通用行解析逻辑
    rows: 迭代器，每项是一个列表/元组
    返回: (added_papers, added_points, db_data_fragment)
    """
    db_data = {"papers": [], "paper_data": []}
    added_papers = 0
    added_points = 0
    
    # 假设第一行是表头，跳过
    it = iter(rows)
    next(it, None) 

    for row_idx, row in enumerate(it, start=2):
        if not row or len(row) < 5: continue
        
        # 1. 基础信息
        doi = str(row[0]).strip()
        raw_article_type = str(row[1]).strip().lower()
        raw_sc_type = str(row[2]).strip().lower()
        formula = str(row[3]).strip()
        structure = str(row[4]).strip() if len(row) > 4 else ""

        article_type = ARTICLE_TYPE_MAP.get(raw_article_type, raw_article_type)
        sc_type = SC_TYPE_MAP.get(raw_sc_type, raw_sc_type)
        element_list = extract_elements(formula)

        try:
            validated_paper = schemas.PaperCreate(
                doi=doi,
                element_symbols=element_list,
                article_type=article_type,
                superconductor_type=sc_type,
                chemical_formula=formula,
                crystal_structure=structure
            )
        except ValidationError as e:
            print(f"❌ 第 {row_idx} 行基本信息校验失败: {e}")
            continue

        # 2. 物理数据
        physical_cols = row[5:]
        current_row_data_points = []
        for i in range(0, len(physical_cols), 5):
            group = physical_cols[i:i+5]
            if len(group) < 2: break
            
            p_val = to_float(group[0])
            tc_val = to_float(group[1])
            if tc_val is None: continue

            try:
                validated_data = schemas.PaperData(
                    pressure=p_val if p_val is not None else 0.0,
                    tc=tc_val,
                    lambda_val=to_float(group[2]) if len(group) > 2 else None,
                    omega_log=to_float(group[3]) if len(group) > 3 else None,
                    n_ef=to_float(group[4]) if len(group) > 4 else None
                )
                current_row_data_points.append(validated_data.model_dump())
            except ValidationError:
                continue

        if not current_row_data_points: continue

        # 3. 组装
        paper_id_temp = added_papers + 1000000 # 临时 ID，仅用于 fragment
        paper_dict = validated_paper.model_dump()
        paper_dict["id_temp"] = paper_id_temp
        paper_dict["element_symbols"] = "-".join(element_list)
        paper_dict["authors"] = json.dumps(["Batch Import"])
        paper_dict["year"] = datetime.now().year
        paper_dict["title"] = f"Imported: {formula}"
        paper_dict["created_at"] = datetime.now().isoformat()
        
        db_data["papers"].append(paper_dict)
        
        for dp in current_row_data_points:
            dp["paper_id_temp"] = paper_id_temp
            db_data["paper_data"].append(dp)
            added_points += 1
        
        added_papers += 1

    return added_papers, added_points, db_data

def process_file(file_content, filename):
    """
    处理上传的文件
    file_content: bytes
    filename: 文件名
    """
    ext = Path(filename).suffix.lower()
    rows = []
    
    if ext == '.xlsx':
        f = io.BytesIO(file_content)
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(row)
    elif ext in ['.csv', '.txt']:
        text = file_content.decode('utf-8-sig')
        f = io.StringIO(text)
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)
    else:
        raise ValueError("不支持的文件格式")
        
    return parse_rows(rows)

if __name__ == "__main__":
    print(CSV_EXAMPLE_TEXT)
    example_path = Path("data/batch_import_example.xlsx")
    if not example_path.exists():
        create_example_xlsx(example_path)

    if len(sys.argv) < 2:
        print("用法: python3 -m backend.merge_csv <input_file> [output_json]")
    else:
        in_path = sys.argv[1]
        out_path = sys.argv[2] if len(sys.argv) > 2 else "data/data_export.json"
        
        with open(in_path, 'rb') as f:
            content = f.read()
        
        added_p, added_d, fragment = process_file(content, in_path)
        
        # 合并逻辑 (CLI 版本依然写到 JSON)
        dest = Path(out_path)
        if dest.exists():
            with open(dest, 'r', encoding='utf-8') as f:
                full_db = json.load(f)
        else:
            full_db = {"papers": [], "paper_data": [], "paper_images": []}
            
        # 转换临时 ID 到真实 ID
        next_p_id = max([p['id'] for p in full_db['papers']], default=0) + 1
        next_d_id = max([d['id'] for d in full_db['paper_data']], default=0) + 1
        
        temp_to_real = {}
        for p in fragment["papers"]:
            old_temp = p.pop("id_temp")
            p["id"] = next_p_id
            temp_to_real[old_temp] = next_p_id
            full_db["papers"].append(p)
            next_p_id += 1
            
        for d in fragment["paper_data"]:
            old_temp = d.pop("paper_id_temp")
            d["id"] = next_d_id
            d["paper_id"] = temp_to_real[old_temp]
            full_db["paper_data"].append(d)
            next_d_id += 1
            
        with open(dest, 'w', encoding='utf-8') as f:
            json.dump(full_db, f, ensure_ascii=False, indent=2)
            
        print(f"✅ 处理完成！新增文献: {added_p}，数据点: {added_d}")
